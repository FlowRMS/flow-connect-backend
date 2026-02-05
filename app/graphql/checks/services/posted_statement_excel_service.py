import io
from datetime import date
from decimal import Decimal

from commons.db.v6.commission import Check
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from app.graphql.checks.strawberry.posted_statement_response import (
    PostedStatementDetailResponse,
    PostedStatementHeaderResponse,
    PostedStatementRepSummaryResponse,
    PostedStatementSummaryResponse,
)


class PostedStatementExcelService:
    def generate_excel(
        self,
        check: Check,
        header: PostedStatementHeaderResponse,
        rep_summaries: list[PostedStatementRepSummaryResponse],
        details: list[PostedStatementDetailResponse],
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Statement")
        ws.title = "Statement"
        summary = self._calculate_summary(check, details)
        self._write_header_section(ws, header)
        self._write_summary_section(ws, summary)
        self._write_rep_summary_section(ws, rep_summaries)
        self._write_detail_section(ws, details, len(rep_summaries))
        self._adjust_column_widths(ws)
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _calculate_summary(
        self,
        check: Check,
        details: list[PostedStatementDetailResponse],
    ) -> PostedStatementSummaryResponse:
        paid = sum(
            (d.commission_received for d in details if d.entity_type == "Invoice"),
            Decimal(0),
        )
        credits = sum(
            (d.commission_received for d in details if d.entity_type == "Credit"),
            Decimal(0),
        )
        expenses = sum(
            (d.commission_received for d in details if d.entity_type == "Adjustment"),
            Decimal(0),
        )
        applied_total = paid + credits + expenses
        expected = sum((d.expected_commission for d in details), Decimal(0))
        adjusted_expected = expected
        balance = expected - applied_total
        return PostedStatementSummaryResponse(
            paid_commissions=paid,
            credits=credits,
            expenses=expenses,
            applied_total=applied_total,
            expected_commission=expected,
            adjusted_expected_commission=adjusted_expected,
            balance=balance,
        )

    def _write_header_section(self, ws, header: PostedStatementHeaderResponse) -> None:  # noqa: ANN001
        header_fill = PatternFill(start_color="FFFF00", fill_type="solid")
        bold = Font(bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        headers = [
            "Post Date",
            "Factory",
            "Check Number",
            "Check Date",
            "Check Commission Month",
            "Commission Amount",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = bold
            cell.border = thin_border
        values = [
            self._format_date(header.post_date),
            header.factory_name,
            header.check_number,
            self._format_date(header.entity_date),
            self._format_date(header.commission_month),
            float(header.commission_amount or 0),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=2, column=col, value=v)
            cell.border = thin_border

    def _write_summary_section(
        self, ws, summary: PostedStatementSummaryResponse
    ) -> None:  # noqa: ANN001
        row = 5
        labels = [
            "Paid Commissions",
            "Credits",
            "Expenses",
            "Applied Total",
            "Expected Commission",
            "Adjusted Expected Commission",
            "Balance",
        ]
        values = [
            summary.paid_commissions,
            summary.credits,
            summary.expenses,
            summary.applied_total,
            summary.expected_commission,
            summary.adjusted_expected_commission,
            summary.balance,
        ]
        for col, (label, val) in enumerate(zip(labels, values, strict=True), 1):
            ws.cell(row=row, column=col, value=label)
            ws.cell(row=row + 1, column=col, value=float(val))

    def _write_rep_summary_section(
        self,
        ws,
        rep_summaries: list[PostedStatementRepSummaryResponse],  # noqa: ANN001
    ) -> None:
        row = 8
        yellow_fill = PatternFill(start_color="FFFF00", fill_type="solid")
        bold = Font(bold=True)
        headers = ["Outside Sales Rep", "Expected Commission", "Commission Received"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = yellow_fill
            cell.font = bold
        for i, rep in enumerate(rep_summaries):
            r = row + 1 + i
            for col, val in enumerate(
                [
                    rep.outside_sales_rep_name,
                    float(rep.expected_commission),
                    float(rep.commission_received),
                ],
                1,
            ):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = yellow_fill

    def _write_detail_section(
        self,
        ws,
        details: list[PostedStatementDetailResponse],
        rep_summary_count: int,
    ) -> None:
        start_row = 8 + rep_summary_count + 2
        headers = [
            "Entity Number",
            "Expected Commission",
            "Commission Received",
            "Outside Sales Rep",
            "Factory",
            "Posted Month",
            "Commission Month",
            "Order Number",
            "Sales Amount",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=col, value=h)
        for i, detail in enumerate(details):
            r = start_row + 1 + i
            vals = [
                detail.entity_number,
                float(detail.expected_commission),
                float(detail.commission_received),
                detail.outside_sales_rep_name,
                detail.factory_name,
                self._format_date(detail.posted_month),
                self._format_date(detail.commission_month),
                detail.order_number or "",
                float(detail.sales_amount),
            ]
            for col, v in enumerate(vals, 1):
                ws.cell(row=r, column=col, value=v)

    def _adjust_column_widths(self, ws) -> None:  # noqa: ANN001
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

    def _format_date(self, d: date | None) -> str:
        return d.strftime("%m/%d/%Y") if d else ""
