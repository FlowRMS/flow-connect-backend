import io
from collections import defaultdict
from datetime import date
from decimal import Decimal
from typing import TypedDict
from uuid import UUID

from commons.auth import AuthInfo
from commons.db.v6.commission import Check, CheckDetail
from openpyxl import Workbook
from openpyxl.styles import Border, Font, PatternFill, Side

from app.graphql.checks.repositories.checks_repository import ChecksRepository
from app.graphql.checks.strawberry.posted_statement_response import (
    PostedStatementDetailResponse,
    PostedStatementHeaderResponse,
    PostedStatementRepSummaryResponse,
    PostedStatementResponse,
    PostedStatementSummaryResponse,
)
from app.graphql.v2.files.services.file_upload_service import FileUploadService


class _RepTotal(TypedDict):
    expected: Decimal
    received: Decimal
    name: str


class _RepSummaryEntry(TypedDict):
    expected: Decimal
    received: Decimal
    name: str
    id: UUID


class PostedStatementService:
    def __init__(  # pyright: ignore[reportMissingSuperCall]
        self,
        checks_repository: ChecksRepository,
        file_upload_service: FileUploadService,
        auth_info: AuthInfo,
    ) -> None:
        self.checks_repository = checks_repository
        self.file_upload_service = file_upload_service
        self.auth_info = auth_info

    async def generate_posted_statement(
        self, check_id: UUID
    ) -> PostedStatementResponse:
        check = await self.checks_repository.find_check_by_id(check_id)
        header = self._build_header(check)
        details = self._build_details(check)
        rep_summaries = self._build_rep_summaries(details)
        excel_bytes = self._generate_excel(check, header, rep_summaries, details)
        upload_result = await self.file_upload_service.upload_file(
            file_content=excel_bytes,
            file_name=f"posted_statement_{check.check_number}.xlsx",
            folder_path=f"statements/{self.auth_info.tenant_id}",
        )
        return PostedStatementResponse(
            header=header,
            rep_summaries=rep_summaries,
            details=details,
            presigned_url=upload_result.presigned_url,
        )

    def _build_header(self, check: Check) -> PostedStatementHeaderResponse:
        return PostedStatementHeaderResponse(
            check_number=check.check_number,
            post_date=check.post_date or check.entity_date,
            entity_date=check.entity_date,
            commission_month=check.commission_month,
            commission_amount=check.entered_commission_amount,
            factory_id=check.factory_id,
            factory_name=check.factory.title,
        )

    def _build_details(self, check: Check) -> list[PostedStatementDetailResponse]:
        details: list[PostedStatementDetailResponse] = []
        for check_detail in check.details:
            details.extend(self._process_check_detail(check, check_detail))
        return details

    def _process_check_detail(
        self, check: Check, check_detail: CheckDetail
    ) -> list[PostedStatementDetailResponse]:
        if check_detail.invoice:
            return self._process_invoice_detail(check, check_detail)
        if check_detail.adjustment:
            return self._process_adjustment_detail(check, check_detail)
        if check_detail.credit:
            return self._process_credit_detail(check, check_detail)
        return []

    def _process_invoice_detail(
        self, check: Check, check_detail: CheckDetail
    ) -> list[PostedStatementDetailResponse]:
        invoice = check_detail.invoice
        if not invoice:
            return []

        order_number = invoice.order.order_number if invoice.order else None
        sales_amount = invoice.balance.total if invoice.balance else Decimal(0)
        total_invoice_commission = (
            invoice.balance.commission if invoice.balance else Decimal(0)
        )

        rep_totals: dict[UUID, _RepTotal] = {}
        for invoice_detail in invoice.details:
            detail_commission = invoice_detail.commission or Decimal(0)
            detail_proportion = (
                detail_commission / total_invoice_commission
                if total_invoice_commission > 0
                else Decimal(0)
            )
            for split_rate in invoice_detail.outside_split_rates:
                rate = (split_rate.split_rate or Decimal(100)) / Decimal(100)
                expected = detail_commission * rate
                received = check_detail.applied_amount * detail_proportion * rate
                rep_id = split_rate.user_id
                if rep_id not in rep_totals:
                    rep_totals[rep_id] = _RepTotal(
                        expected=Decimal(0),
                        received=Decimal(0),
                        name=split_rate.user.full_name,
                    )
                rep_totals[rep_id]["expected"] += expected
                rep_totals[rep_id]["received"] += received

        return [
            PostedStatementDetailResponse(
                entity_number=invoice.invoice_number,
                entity_type="Invoice",
                expected_commission=data["expected"],
                commission_received=data["received"].quantize(Decimal("0.01")),
                outside_sales_rep_id=rep_id,
                outside_sales_rep_name=data["name"],
                factory_name=check.factory.title,
                posted_month=check.post_date,
                commission_month=check.commission_month,
                order_number=order_number,
                sales_amount=sales_amount,
            )
            for rep_id, data in rep_totals.items()
        ]

    def _process_adjustment_detail(
        self, check: Check, check_detail: CheckDetail
    ) -> list[PostedStatementDetailResponse]:
        details: list[PostedStatementDetailResponse] = []
        adjustment = check_detail.adjustment
        if not adjustment:
            return details

        for split_rate in adjustment.split_rates:
            rate = (split_rate.split_rate or Decimal(100)) / Decimal(100)
            expected = adjustment.amount * rate
            received = check_detail.applied_amount * rate
            details.append(
                PostedStatementDetailResponse(
                    entity_number=adjustment.adjustment_number,
                    entity_type="Adjustment",
                    expected_commission=expected,
                    commission_received=received,
                    outside_sales_rep_id=split_rate.user_id,
                    outside_sales_rep_name=split_rate.user.full_name,
                    factory_name=check.factory.title,
                    posted_month=check.post_date,
                    commission_month=check.commission_month,
                    order_number=None,
                    sales_amount=adjustment.amount,
                )
            )
        return details

    def _process_credit_detail(
        self, check: Check, check_detail: CheckDetail
    ) -> list[PostedStatementDetailResponse]:
        credit = check_detail.credit
        if not credit:
            return []

        order_number = credit.order.order_number if credit.order else None
        sales_amount = credit.balance.total if credit.balance else Decimal(0)
        total_credit_commission = (
            credit.balance.commission if credit.balance else Decimal(0)
        )

        rep_totals: dict[UUID, _RepTotal] = {}
        for credit_detail in credit.details:
            detail_commission = credit_detail.commission or Decimal(0)
            detail_proportion = (
                detail_commission / total_credit_commission
                if total_credit_commission > 0
                else Decimal(0)
            )
            for split_rate in credit_detail.outside_split_rates:
                rate = (split_rate.split_rate or Decimal(100)) / Decimal(100)
                expected = detail_commission * rate
                received = check_detail.applied_amount * detail_proportion * rate
                rep_id = split_rate.user_id
                if rep_id not in rep_totals:
                    rep_totals[rep_id] = _RepTotal(
                        expected=Decimal(0),
                        received=Decimal(0),
                        name=split_rate.user.full_name,
                    )
                rep_totals[rep_id]["expected"] += expected
                rep_totals[rep_id]["received"] += received

        return [
            PostedStatementDetailResponse(
                entity_number=credit.credit_number,
                entity_type="Credit",
                expected_commission=Decimal("-1") * data["expected"],
                commission_received=(Decimal("-1") * data["received"]).quantize(
                    Decimal("0.01")
                ),
                outside_sales_rep_id=rep_id,
                outside_sales_rep_name=data["name"],
                factory_name=check.factory.title,
                posted_month=check.post_date,
                commission_month=check.commission_month,
                order_number=order_number,
                sales_amount=sales_amount,
            )
            for rep_id, data in rep_totals.items()
        ]

    def _build_rep_summaries(
        self, details: list[PostedStatementDetailResponse]
    ) -> list[PostedStatementRepSummaryResponse]:
        summaries: dict[UUID, _RepSummaryEntry] = defaultdict(
            lambda: _RepSummaryEntry(
                expected=Decimal(0),
                received=Decimal(0),
                name="",
                id=UUID(int=0),
            )
        )
        for detail in details:
            rep_id = detail.outside_sales_rep_id
            summaries[rep_id]["expected"] += detail.expected_commission
            summaries[rep_id]["received"] += detail.commission_received
            summaries[rep_id]["name"] = detail.outside_sales_rep_name
            summaries[rep_id]["id"] = rep_id
        return [
            PostedStatementRepSummaryResponse(
                outside_sales_rep_id=data["id"],
                outside_sales_rep_name=data["name"],
                expected_commission=data["expected"],
                commission_received=data["received"],
            )
            for data in summaries.values()
        ]

    def _calculate_summary(
        self,
        check: Check,
        details: list[PostedStatementDetailResponse],
    ) -> PostedStatementSummaryResponse:
        paid = sum(
            (d.commission_received for d in details if d.entity_type == "Invoice"),
            Decimal(0),
        )
        credits = sum(
            (d.commission_received for d in details if d.entity_type == "Credit"),
            Decimal(0),
        )
        expenses = sum(
            (d.commission_received for d in details if d.entity_type == "Adjustment"),
            Decimal(0),
        )
        applied_total = paid + credits + expenses
        expected = sum((d.expected_commission for d in details), Decimal(0))
        adjusted_expected = expected
        balance = expected - applied_total
        return PostedStatementSummaryResponse(
            paid_commissions=paid,
            credits=credits,
            expenses=expenses,
            applied_total=applied_total,
            expected_commission=expected,
            adjusted_expected_commission=adjusted_expected,
            balance=balance,
        )

    def _generate_excel(
        self,
        check: Check,
        header: PostedStatementHeaderResponse,
        rep_summaries: list[PostedStatementRepSummaryResponse],
        details: list[PostedStatementDetailResponse],
    ) -> bytes:
        wb = Workbook()
        ws = wb.active
        if ws is None:
            ws = wb.create_sheet("Statement")
        ws.title = "Statement"
        summary = self._calculate_summary(check, details)
        self._write_header_section(ws, header)
        self._write_summary_section(ws, summary)
        self._write_rep_summary_section(ws, rep_summaries)
        self._write_detail_section(ws, details, len(rep_summaries))
        self._adjust_column_widths(ws)
        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def _write_header_section(self, ws, header: PostedStatementHeaderResponse) -> None:  # noqa: ANN001
        header_fill = PatternFill(start_color="FFFF00", fill_type="solid")
        bold = Font(bold=True)
        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        headers = [
            "Post Date",
            "Factory",
            "Check Number",
            "Check Date",
            "Check Commission Month",
            "Commission Amount",
        ]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = bold
            cell.border = thin_border
        values = [
            self._format_date(header.post_date),
            header.factory_name,
            header.check_number,
            self._format_date(header.entity_date),
            self._format_date(header.commission_month),
            float(header.commission_amount or 0),
        ]
        for col, v in enumerate(values, 1):
            cell = ws.cell(row=2, column=col, value=v)
            cell.border = thin_border

    def _write_summary_section(
        self, ws, summary: PostedStatementSummaryResponse
    ) -> None:  # noqa: ANN001
        row = 5
        labels = [
            "Paid Commissions",
            "Credits",
            "Expenses",
            "Applied Total",
            "Expected Commission",
            "Adjusted Expected Commission",
            "Balance",
        ]
        values = [
            summary.paid_commissions,
            summary.credits,
            summary.expenses,
            summary.applied_total,
            summary.expected_commission,
            summary.adjusted_expected_commission,
            summary.balance,
        ]
        for col, (label, val) in enumerate(zip(labels, values, strict=True), 1):
            ws.cell(row=row, column=col, value=label)
            ws.cell(row=row + 1, column=col, value=float(val))

    def _write_rep_summary_section(
        self,
        ws,
        rep_summaries: list[PostedStatementRepSummaryResponse],  # noqa: ANN001
    ) -> None:
        row = 8
        yellow_fill = PatternFill(start_color="FFFF00", fill_type="solid")
        bold = Font(bold=True)
        headers = ["Outside Sales Rep", "Expected Commission", "Commission Received"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.fill = yellow_fill
            cell.font = bold
        for i, rep in enumerate(rep_summaries):
            r = row + 1 + i
            for col, val in enumerate(
                [
                    rep.outside_sales_rep_name,
                    float(rep.expected_commission),
                    float(rep.commission_received),
                ],
                1,
            ):
                cell = ws.cell(row=r, column=col, value=val)
                cell.fill = yellow_fill

    def _write_detail_section(
        self,
        ws,
        details: list[PostedStatementDetailResponse],
        rep_summary_count: int,
    ) -> None:
        start_row = 8 + rep_summary_count + 2
        headers = [
            "Entity Number",
            "Expected Commission",
            "Commission Received",
            "Outside Sales Rep",
            "Factory",
            "Posted Month",
            "Commission Month",
            "Order Number",
            "Sales Amount",
        ]
        for col, h in enumerate(headers, 1):
            ws.cell(row=start_row, column=col, value=h)
        for i, detail in enumerate(details):
            r = start_row + 1 + i
            vals = [
                detail.entity_number,
                float(detail.expected_commission),
                float(detail.commission_received),
                detail.outside_sales_rep_name,
                detail.factory_name,
                self._format_date(detail.posted_month),
                self._format_date(detail.commission_month),
                detail.order_number or "",
                float(detail.sales_amount),
            ]
            for col, v in enumerate(vals, 1):
                ws.cell(row=r, column=col, value=v)

    def _adjust_column_widths(self, ws) -> None:  # noqa: ANN001
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except Exception:
                    pass
            ws.column_dimensions[column_letter].width = max_length + 2

    def _format_date(self, d: date | None) -> str:
        return d.strftime("%m/%d/%Y") if d else ""
