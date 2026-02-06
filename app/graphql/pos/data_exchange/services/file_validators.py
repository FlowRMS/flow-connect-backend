import csv
import io

from app.graphql.pos.data_exchange.exceptions import InvalidFileTypeError

ALLOWED_EXTENSIONS = {"csv", "xls", "xlsx"}


def validate_file_type(file_name: str) -> str:
    extension = file_name.rsplit(".", 1)[-1].lower() if "." in file_name else ""
    if extension not in ALLOWED_EXTENSIONS:
        raise InvalidFileTypeError(
            f"Invalid file type: {extension}. Allowed: {', '.join(ALLOWED_EXTENSIONS)}"
        )
    return extension


def count_rows(file_content: bytes, file_type: str) -> int:
    if file_type == "csv":
        return _count_csv_rows(file_content)
    if file_type == "xls":
        return _count_xls_rows(file_content)
    if file_type == "xlsx":
        return _count_xlsx_rows(file_content)
    return 0


def _count_csv_rows(file_content: bytes) -> int:
    text = file_content.decode("utf-8")
    reader = csv.reader(io.StringIO(text))
    rows = list(reader)
    return max(0, len(rows) - 1)


# noinspection PyBroadException
# User files can fail in unpredictable ways; return 0 rather than crash
def _count_xls_rows(file_content: bytes) -> int:
    try:
        import xlrd

        workbook = xlrd.open_workbook(file_contents=file_content)
        sheet = workbook.sheet_by_index(0)
        return max(0, sheet.nrows - 1)
    except Exception:
        return 0


# noinspection PyBroadException
# User files can fail in unpredictable ways; return 0 rather than crash
def _count_xlsx_rows(file_content: bytes) -> int:
    try:
        from openpyxl import load_workbook

        workbook = load_workbook(filename=io.BytesIO(file_content), read_only=True)
        sheet = workbook.active
        if sheet is None:
            return 0
        row_count = sum(
            1 for row in sheet.iter_rows() if any(cell.value for cell in row)
        )
        return max(0, row_count - 1)
    except Exception:
        return 0
