import csv
import io
from dataclasses import dataclass
from typing import Any

from commons.s3.service import S3Service

from app.graphql.pos.field_map.models.field_map import FieldMap


class UnsupportedFileTypeError(Exception):
    pass


@dataclass
class FileRow:
    row_number: int
    data: dict[str, Any]


class FileReaderService:
    SUPPORTED_TYPES = {"csv", "xls", "xlsx"}

    def __init__(self, s3_service: S3Service) -> None:
        self.s3_service = s3_service

    async def read_file(
        self,
        s3_key: str,
        file_type: str,
        field_map: FieldMap | None = None,
    ) -> list[FileRow]:
        if file_type not in self.SUPPORTED_TYPES:
            raise UnsupportedFileTypeError(
                f"Unsupported file type: {file_type}. "
                f"Supported: {', '.join(self.SUPPORTED_TYPES)}"
            )

        file_obj = await self.s3_service.download(key=s3_key)
        content = file_obj.read()

        if file_type == "csv":
            rows = self._parse_csv(content)
        elif file_type == "xls":
            rows = self._parse_xls(content)
        else:
            rows = self._parse_xlsx(content)

        if field_map:
            rows = self._apply_field_mapping(rows, field_map)

        return rows

    @staticmethod
    def _parse_csv(content: bytes) -> list[FileRow]:
        text = content.decode("utf-8")
        reader = csv.DictReader(io.StringIO(text))
        return [
            FileRow(row_number=i + 2, data=dict(row)) for i, row in enumerate(reader)
        ]

    @staticmethod
    def _parse_xlsx(content: bytes) -> list[FileRow]:
        from openpyxl import load_workbook

        workbook = load_workbook(filename=io.BytesIO(content), read_only=True)
        sheet = workbook.active
        if sheet is None:
            return []

        rows_iter = sheet.iter_rows(values_only=True)
        headers = next(rows_iter, None)
        if headers is None:
            return []

        header_list = [str(h) if h is not None else "" for h in headers]
        result: list[FileRow] = []
        for i, row in enumerate(rows_iter):
            if not any(cell is not None for cell in row):
                continue
            data = {
                header_list[j]: row[j] for j in range(len(header_list)) if j < len(row)
            }
            result.append(FileRow(row_number=i + 2, data=data))
        return result

    @staticmethod
    def _parse_xls(content: bytes) -> list[FileRow]:
        import xlrd

        workbook = xlrd.open_workbook(file_contents=content)
        sheet = workbook.sheet_by_index(0)
        if sheet.nrows == 0:
            return []

        headers = [str(sheet.cell_value(0, col)) for col in range(sheet.ncols)]
        result: list[FileRow] = []
        for row_idx in range(1, sheet.nrows):
            data = {
                headers[col]: sheet.cell_value(row_idx, col)
                for col in range(sheet.ncols)
            }
            result.append(FileRow(row_number=row_idx + 1, data=data))
        return result

    def _apply_field_mapping(
        self,
        rows: list[FileRow],
        field_map: FieldMap,
    ) -> list[FileRow]:
        column_mapping = self._build_column_mapping(field_map)
        if not column_mapping:
            return rows

        result: list[FileRow] = []
        for row in rows:
            mapped_data: dict[str, Any] = {}
            for col_name, value in row.data.items():
                if col_name in column_mapping:
                    mapped_data[column_mapping[col_name]] = value
                else:
                    mapped_data[col_name] = value
            result.append(FileRow(row_number=row.row_number, data=mapped_data))
        return result

    @staticmethod
    def _build_column_mapping(field_map: FieldMap) -> dict[str, str]:
        mapping: dict[str, str] = {}
        for field in field_map.fields:
            if field.organization_field_name:
                mapping[field.organization_field_name] = field.standard_field_key
        return mapping
